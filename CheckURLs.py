import openpyxl
import requests
import time

# Path to the excel file 
excelFilePath = "C:\\BRG\\URL.xlsx"

# Name of excel sheet tab
worksheetName = "tocheck"

# Load the excel file
workbook = openpyxl.load_workbook(excelFilePath)

# Select the excel source tab 
worksheet = workbook[worksheetName]

# Create a new excel tab to put result , existing tab renamed
checkedWorksheetName = "checked"
checkedWorksheet = workbook.create_sheet(title=checkedWorksheetName)

# Parcourir chaque ligne de l'onglet et vérifier les URL
for row in worksheet.iter_rows(min_row=2, values_only=True):  # Commence à la deuxième ligne pour ignorer les en-têtes
    url = row[1]  # Remplacez l'indice par la colonne souhaitée (0-based index)
    start_time = time.time()
    response = requests.get(url)
    end_time = time.time()

    result = "Accessible" if 200 <= response.status_code < 400 else "Erreur "+str(response.status_code)

    load_time=end_time - start_time
    checkedWorksheet.append([url, resultf"{load_time:.4f} secondes"]])

# Sauvegarder les modifications dans le fichier Excel
workbook.save(excelFilePath)

print(f"Résultats stockés dans l'onglet '{checkedWorksheetName}' du fichier '{excelFilePath}'.")
